import os
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet


class Common:
    def __init__(self):
        self.RESOURCES_DIR = Path(__file__).parent
        self.DATA_STORE_FILE = os.path.join(self.RESOURCES_DIR, "test_data", "DataStore.xlsx")
        self.TEST_CASES_FILE = os.path.join(self.RESOURCES_DIR, "test_data", "TestCases.xlsx")
        self.LANGUAGES_FILE = os.path.join(self.RESOURCES_DIR, 'test_data', 'references', '_list.csv')
        self.HTML_TAGS = ['<br>', '<p>', '<span>']
        self.PATTERN = "_pattern"
        self.CLEAN_CHAR = '_'
        self.REPORT_FILE = os.path.join(self.RESOURCES_DIR.parent.parent, 'results', 'output_backup.xml')
        self.REPORT_FILE2 = os.path.join(self.RESOURCES_DIR.parent.parent, 'results', 'output.xml')

    def clean_up_text(self, txt: str) -> str:
        res = ""
        for t in txt:
             res+= self.CLEAN_CHAR if t.isspace() or not (t.isalnum()) else t
        return res

    def get_position_by_name_and_value(self, sh: Worksheet, field_name: str, field_value: str, contains_name=True) -> (int, int):
        r = 0
        c = self.get_column_by_name(sh, field_name, contains_name)
        for row in range(2, sh.max_row):
            cv = sh.cell(row,c).value
            if (contains_name and cv in field_value) or (not contains_name and cv == field_value):
                    r = row
                    break
        return r, c

    def get_column_by_name(self, sh: Worksheet, field_name, contains_name=True) -> int:
        c = 0
        for col in range(1, sh.max_column+1):
            cv = sh.cell(1,col).value
            if (contains_name and field_name in cv) or (not contains_name and cv == field_name):
                    c = col
                    break
        return c

# c = Common()
# print(c.clean_up_text("154as.>:{]'\"Z*-+A"))
