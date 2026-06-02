"""
Typofix Helper Functions for Playwright Tests

Provides utilities for:
- Excel workbook management (TestCases.xlsx)
- Hyperlink extraction and manipulation
- Test result recording
- Web element interaction helpers
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class TypofixHelpers:
    """Helper class for Typofix test operations"""

    def __init__(self):
        """Initialize helper with paths and configuration"""
        self.RESOURCES_DIR = Path(__file__).parent.parent
        self.TEST_CASES_FILE = os.path.join(self.RESOURCES_DIR, "test_data", "TestCases.xlsx")
        self.TEST_CASES_WB = load_workbook(self.TEST_CASES_FILE)
        self.LANGUAGES_FILE = os.path.join(self.RESOURCES_DIR, 'test_data', 'references', '_list.csv')
        self.TEST_RESULTS_FIELDS = ("TEST_RESULT", "REAL", "DETAILS", "TIMESTAMP", "SCREENSHOT")
        self.HTML_TAGS = ['<br>', '<p>', '<span>']
        self.PATTERN = "_pattern"
        self.CLEAN_CHAR = '_'

    def get_hyperlink_by_link_name(self, column_name: str, value: str) -> str:
        """
        Extract hyperlink from Excel cell by column name and value.

        Args:
            column_name: Name of the column to search in
            value: Value to match in the column

        Returns:
            URL string from the hyperlink
        """
        sh = self.TEST_CASES_WB.worksheets[0]
        r, c = self._get_position_by_name_and_value(sh, column_name, value)
        link = sh.cell(r, c).hyperlink
        return str(link.target) if link else ""

    def create_new_excel_list_in_excel(self) -> str:
        """
        Create a new worksheet from the pattern template.

        Returns:
            Name of the newly created worksheet
        """
        first = self.TEST_CASES_WB.copy_worksheet(self.TEST_CASES_WB[self.PATTERN])
        first.title = "tc_A"
        self.TEST_CASES_WB.move_sheet(first, offset=-(len(self.TEST_CASES_WB.worksheets) - 1))
        print(f"Selected target worksheet {first.title}")
        return first.title

    def add_new_test_cases_to_excel(
        self,
        excel_list: str,
        id: str,
        name: str,
        url_detail: str,
        languages: list,
        befores: list,
        afters: list
    ) -> None:
        """
        Add new test case rows to Excel worksheet.

        Args:
            excel_list: Target worksheet name
            id: Rule ID
            name: Rule name
            url_detail: URL to detailed rule page
            languages: List of language codes
            befores: List of before (input) examples
            afters: List of after (expected) examples
        """
        ws = self.TEST_CASES_WB[excel_list]
        rows = ws.max_row
        for i, language in enumerate(languages):
            test_name = self._clean_up_text(id + self.CLEAN_CHAR + name + self.CLEAN_CHAR + language.strip())
            ws.cell(row=rows + i, column=1, value=test_name)
            self._insert_excel_hyperlink(
                ws.cell(row=rows + i, column=4),
                id + " - " + name.strip(),
                url_detail.strip()
            )
            ws.cell(row=rows + i, column=5, value=language.strip())
            ws.cell(row=rows + i, column=6, value=befores[i].strip())
            ws.cell(row=rows + i, column=7, value=afters[i].strip())

    def add_results_to_excel(self, test_name: str, *f_values) -> str:
        """
        Add test results to the Excel file.

        Args:
            test_name: Name of the test case
            *f_values: Test result field values (TEST_RESULT, REAL, DETAILS, TIMESTAMP, SCREENSHOT)

        Returns:
            Error message if test case not found, empty string otherwise
        """
        errors = ""
        sh = self.TEST_CASES_WB.worksheets[0]
        row, x = self._get_position_by_name_and_value(sh, "Test Cases", test_name, False)
        if row == 0:
            errors = f"Test Case {test_name} not found"
        else:
            for i, field in enumerate(self.TEST_RESULTS_FIELDS):
                print(self.TEST_RESULTS_FIELDS[i], f_values[i])
                col = self._get_column_by_name(sh, field)
                sh.cell(row, col).value = f_values[i]
        return errors

    def save_test_case_excel(self) -> None:
        """Save all changes to the TestCases.xlsx file"""
        self.TEST_CASES_WB.save(self.TEST_CASES_FILE)

    def _get_position_by_name_and_value(
        self,
        sh: Worksheet,
        field_name: str,
        field_value: str,
        contains_name: bool = True
    ) -> tuple:
        """
        Find cell position by field name and value.

        Args:
            sh: Worksheet to search in
            field_name: Column header name
            field_value: Value to match
            contains_name: If True, use partial matching; if False, use exact matching

        Returns:
            Tuple of (row, column) indices
        """
        r = 0
        c = self._get_column_by_name(sh, field_name, True)
        for row in range(2, sh.max_row):
            cv = sh.cell(row, c).value
            if cv and ((contains_name and cv in field_value) or (not contains_name and cv == field_value)):
                r = row
                break
        return r, c

    @staticmethod
    def _get_column_by_name(sh: Worksheet, field_name: str, contains_name: bool = True) -> int:
        """
        Find column index by field name.

        Args:
            sh: Worksheet to search in
            field_name: Column header name to find
            contains_name: If True, use partial matching; if False, use exact matching

        Returns:
            Column index (1-based)
        """
        c = 0
        for col in range(1, sh.max_column + 1):
            cv = sh.cell(1, col).value
            if cv and ((contains_name and field_name in cv) or (not contains_name and cv == field_name)):
                c = col
                break
        return c

    def _insert_excel_hyperlink(self, cell: Cell, name: str, link: str) -> None:
        """
        Insert a hyperlink into an Excel cell.

        Args:
            cell: Target cell
            name: Display name for the hyperlink
            link: URL to link to
        """
        cell.value = name
        cell.hyperlink = self._customize_url(link)

    @staticmethod
    def _customize_url(url: str, pattern_name: str = 'detail') -> str:
        """
        Customize URL for Excel hyperlink.

        Args:
            url: Original URL
            pattern_name: Pattern type ('detail' removes query parameters)

        Returns:
            Customized URL
        """
        if pattern_name == 'detail' and '=' in url:
            url = url[:url.index('=')]
        return url

    def _clean_up_text(self, txt: str) -> str:
        """
        Clean up text for use as test case name.

        Replaces spaces and special characters with underscores.

        Args:
            txt: Input text

        Returns:
            Cleaned text
        """
        res = ""
        for t in txt:
            res += self.CLEAN_CHAR if t.isspace() or not t.isalnum() else t
        return res
