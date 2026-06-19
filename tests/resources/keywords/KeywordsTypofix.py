import os
import datetime
from enum import Enum
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText,TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.comments import Comment

class KeywordsTypofix(object):

    def __init__(self):
        self.XLS_GREEN = InlineFont(color='00008000')
        self.XLS_RED = InlineFont(color='00FF0000')
        class TResult(Enum):
            SKIPPED = "SKIPPED"
            PASSED = "PASSED"
            FAILED = "FAILED"
        self.TR = TResult
        self.RESOURCES_DIR = Path(__file__).parent.parent
        self.TEST_CASES_FILE = os.path.join(self.RESOURCES_DIR, "test_data", "TestCases.xlsx")
        self.TEST_CASES_WB = load_workbook(self.TEST_CASES_FILE)
        self.TEST_CASES_MISSING = self.TEST_CASES_WB["tc_no_examples"]
        self.MISSING_LL = 2
        self.LL = 2
        self.LANGUAGES_FILE = os.path.join(self.RESOURCES_DIR, 'test_data', 'references', '_list.csv')
        self.TEST_RESULTS_FIELDS = ("TEST_RESULT", "REAL", "DETAILS", "TIMESTAMP")
        self.HTML_TAGS = ['<br>', '<p>', '<span>']
        self.PATTERN = "_pattern"
        self.SHEET_NAME = "TC"
        self.CLEAN_CHAR = '_'
        self.URL_DETAIL = 'https://typofix.slonline.sk/admin/rules/SLONline-Typofix-Model-Rule/EditForm/field/SLONline-Typofix-Model-Rule/item/'
        self.FILE_LOG = os.path.join(self.RESOURCES_DIR, 'log.txt')
        self.XLS_GREEN = InlineFont(color='00008000')
        self.XLS_RED = InlineFont(color='00FF0000')


    @staticmethod
    def get_columns_from_data(data: list ):
        ids  = [data_item[0]  for data_item  in  data]
        names = [data_item[1]  for data_item  in  data]
        descriptions = [data_item[2]  for data_item  in  data]
        tags  = [data_item[3]  for data_item  in  data]
        expected_languages = [data_item[4]  for data_item  in  data]
        return ids, names,  descriptions, tags, expected_languages

    def typofix_file_log(self, line = "", is_new=False):
        if is_new:
            open(self.FILE_LOG, 'w').close()
        if line != "":
            with open(self.FILE_LOG, 'a') as file:
                file.write(line + "\n")

    def put_note_to_excel(self, cnt_ok: int, cnt_total: int, ws_name="tc_no_examples") -> None:
        note = f"Last execution at {datetime.datetime.now()}. Reported records: {cnt_ok}.  Total count: {cnt_total}"
        ws = self.TEST_CASES_WB[ws_name]
        ws["A1"].comment = Comment(note, "TypeFix test automation")

    @staticmethod
    def build_before_after_for_languages(data: list, languages_examples: list[str], expected_languages: str):
        even = []
        odd = []
        before = []
        after = []
        final_languages = []
        expected_languages_list = expected_languages.split(", ")
        for i in range(0, len(data), 2):
            data[i] = "\n".join(data[i])
            data[i + 1] = "\n".join(data[i+1])
            odd.append(data[i])
            even.append(data[i + 1])
        for i, language_example in enumerate(languages_examples):
            if language_example in expected_languages_list:
                before.append(odd[i])
                after.append(even[i])
                final_languages.append(language_example)
        return final_languages, before, after

    def get_hyperlink_by_link_name(self, column_name: str, value) -> str:
        sh = self.TEST_CASES_WB.worksheets[0]
        r, c = self._get_position_by_name_and_value(sh, column_name, value)
        link = sh.cell(r, c).hyperlink
        return str(link.target)

    def create_new_excel_list_in_excel(self, title= None, use_pattern= True) -> str:
        if use_pattern:
            first = self.TEST_CASES_WB.copy_worksheet(self.TEST_CASES_WB[self.PATTERN])
            self.TEST_CASES_WB.move_sheet(first, offset=- (len(self.TEST_CASES_WB.worksheets) - 1))
        else:
            first = self.TEST_CASES_WB.create_sheet(title,0)
        first.title = self.SHEET_NAME if title is None else title
        return first.title

    def clean_missing_excel_list(self):
        for row in self.TEST_CASES_MISSING['A2:Z9000']:
            for cell in row:
                cell.value = None

    def add_missing_examples_to_excel(self, id: str, name: str, description: str, tag: str, expected_languages: str):
        languages = expected_languages.split(", ")
        for language in languages:
            self.TEST_CASES_MISSING.cell(row=self.MISSING_LL, column=1, value=self._clean_up_text(id + self.CLEAN_CHAR + name + self.CLEAN_CHAR + language.strip()))
            self.TEST_CASES_MISSING.cell(row=self.MISSING_LL, column=2, value=description)
            self.TEST_CASES_MISSING.cell(row=self.MISSING_LL, column=3, value=self._clean_up_text(tag))
            self.TEST_CASES_MISSING.cell(row=self.MISSING_LL, column=4, value=id)
            self._insert_excel_hyperlink(self.TEST_CASES_MISSING.cell(row=self.MISSING_LL, column=5), id + self.CLEAN_CHAR + name.strip(), id=id)
            self.TEST_CASES_MISSING.cell(row=self.MISSING_LL, column=6, value=language.strip())
            self.MISSING_LL += 1

    def add_new_test_cases_to_excel(self, excel_list: str, id: str, name: str, description: str, tag: str,languages: list[str], befores: list, afters: list):
        ws = self.TEST_CASES_WB[excel_list]
        for i, language in enumerate(languages):
            ws.cell(row=self.LL, column=1, value=self._clean_up_text(id + self.CLEAN_CHAR + name + self.CLEAN_CHAR + language.strip()))
            ws.cell(row=self.LL, column=2, value=description)
            ws.cell(row=self.LL, column=3, value=self._clean_up_text(tag))
            ws.cell(row=self.LL, column=4, value=id)
            self._insert_excel_hyperlink(ws.cell(row=self.LL, column=5), id + " - " + name.strip(), id=id)
            ws.cell(row=self.LL, column=6, value=language.strip())
            ws.cell(row=self.LL, column=7, value=befores[i])
            ws.cell(row=self.LL, column=8, value=self.format_nbspace_character(afters[i]))
            self.LL += 1

    def add_table_to_excel(self, *f_values):
        sh = self.TEST_CASES_WB.worksheets[0]
        for r_id, row in  enumerate(f_values):
            for c_id, vx in enumerate(row):
                v =  "\n".join(vx) if isinstance(vx, list) else str(vx)
                sh.cell(row=r_id + 1, column=c_id + 1, value=v)

    def add_results_to_excel(self, after:str, test_name:str, *f_values):
        errors = ""
        sh = self.TEST_CASES_WB.worksheets[0]
        row, x = self._get_position_by_name_and_value(sh, "Test Cases", test_name, False)
        if row == 0:
            errors = f"Test Case {test_name} not found"
        else:
            for i, field in enumerate(self.TEST_RESULTS_FIELDS):
                col = self._get_column_by_name(sh, field)
                if field == "REAL":
                    self._color_cell_text(sh.cell(row, col), after, f_values[i])
                else:
                    sh.cell(row, col).value = f_values[i]
        return errors

    def save_test_case_excel(self) -> None:
        self.TEST_CASES_WB.save(self.TEST_CASES_FILE)

    def get_detail_link(self, id: str) -> str:
        return self.URL_DETAIL + id + "/edit?Root_LanguageExamples#Root_LanguageExamples"

    @staticmethod
    def typofix_split_string(s:str) -> list[str]:
        return [ss.strip() for ss in s.split(",")]

    def format_nbspace_character_for_list (self, txt_list: list) -> list:
        return [self.format_nbspace_character(txt) for txt in txt_list]


    def format_nbspace_character (self, txt: str) -> str:
        if not txt: return ""
        pairs = {chr(160): "&nbSpace;", "\u2009": "&nbSpace;", "\n": "&endLine;", "\t": "&tab;"}
        for pk in pairs.keys():
            txt = txt.replace(pk, pairs[pk])
        return  txt

    def _get_position_by_name_and_value(self, sh: Worksheet, field_name: str, field_value: str, contains_name=True) -> tuple[int, int]:
        r = 0
        c = self._get_column_by_name(sh, field_name, True)
        for row in range(2, sh.max_row + 1):
            cv = sh.cell(row, c).value
            if (contains_name and cv in field_value) or (not contains_name and cv == field_value):
                r = row
                break
        return r, c

    @staticmethod
    def _get_column_by_name(sh: Worksheet, field_name, contains_name=True) -> int:
        c = 0
        for col in range(1, sh.max_column + 1):
            cv = sh.cell(1, col).value
            if (contains_name and field_name in cv) or (not contains_name and cv == field_name):
                c = col
                break
        return c

    def _insert_excel_hyperlink(self, c: Cell, name: str, id: str) -> None:
        c.value = name
        c.hyperlink = self.get_detail_link(id)

    def _clean_up_text(self, txt: str) -> str:
        res = ""
        for t in txt:
            res += t if t.isalnum() else ' '
        res = ' '.join(res.split())
        res = res.replace(' ', self.CLEAN_CHAR)
        return res

    def assert_custom_typofix (self, after: str, real: str ):
        if  after.strip() == "": return  self.TR.SKIPPED.value, "Empty AFTER field"
        result = self.TR.PASSED.value if after == real  else self.TR.FAILED.value
        details = "" if after == real  else f"'{after}' is not equal to '{real}'"
        return  result,  details

    def _color_cell_text(self, cell: Cell, after: str, real: str) -> None:
        new_real = ""
        ix = 0
        for ix, e_char in enumerate(list(after)):
            if e_char != (list(real))[ix]:  break
            new_real += e_char
        cell.value = real if new_real == real else CellRichText([TextBlock(self.XLS_GREEN, new_real), TextBlock(self.XLS_RED,real[ix:])])

# tp = KeywordsTypofix()
# sh_name = tp.create_new_excel_list_in_excel("TEST", False)
# cll = tp.TEST_CASES_WB[sh_name]["A1"]
# tp.color_cell_text(cll,"ABCDEF", "ABCxDEF")
# cll = tp.TEST_CASES_WB[sh_name]["A2"]
# tp.color_cell_text(cll,"ABCDEF", "ABCDEF")
# tp.save_test_case_excel()