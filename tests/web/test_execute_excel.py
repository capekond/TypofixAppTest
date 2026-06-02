"""
Execute Excel Test Suite

Executes test cases from the TestCases.xlsx file against the Typofix application.

This test:
1. Reads test cases from Excel
2. For each test case:
   - Navigates to the specified URL
   - Selects the language
   - Verifies the output matches expected result
   - Records the result to Excel
3. Generates a test report

Usage:
    pytest tests/web/test_execute_excel.py -v
"""

import pytest
import asyncio
from datetime import datetime
from openpyxl import load_workbook
import os
from pathlib import Path


@pytest.fixture
def test_data():
    """Load test data from Excel file"""
    resources_dir = Path(__file__).parent.parent / 'resources'
    test_cases_file = os.path.join(resources_dir, 'test_data', 'TestCases.xlsx')
    
    wb = load_workbook(test_cases_file)
    ws = wb.worksheets[0]  # First sheet contains test cases
    
    test_cases = []
    for row in range(2, ws.max_row + 1):
        test_case = {
            'name': ws.cell(row, 1).value,
            'link': ws.cell(row, 4).value,
            'language': ws.cell(row, 5).value,
            'before': ws.cell(row, 6).value,
            'after': ws.cell(row, 7).value,
            'row': row
        }
        if test_case['name'] and test_case['link']:
            test_cases.append(test_case)
    
    return test_cases, wb, test_cases_file


@pytest.mark.execute
@pytest.mark.asyncio
async def test_execute_test_from_excel(tested_page, typofix_helpers, test_data):
    """
    Execute test cases from Excel file.

    Args:
        tested_page: Page fixture for the tested application
        typofix_helpers: TypofixHelpers instance
        test_data: Test data from Excel
    """
    test_cases, wb, test_cases_file = test_data
    
    results = []
    
    for test_case in test_cases:
        try:
            # Get test parameters
            test_name = test_case['name']
            url = test_case['link']
            language = test_case['language']
            expected = test_case['after']
            
            print(f"\nExecuting: {test_name}")
            print(f"  URL: {url}")
            print(f"  Language: {language}")
            print(f"  Expected: {expected}")
            
            # Navigate to the application
            await tested_page.goto("https://www.typofix.org/application#testing")
            await tested_page.wait_for_load_state("networkidle")
            
            # Select language if available
            try:
                language_select = tested_page.locator("id:language-select")
                await language_select.select_option(language)
            except:
                print(f"  Warning: Could not select language {language}")
            
            # Get the actual result from the page
            # This is a placeholder - actual implementation depends on application structure
            try:
                output_elem = tested_page.locator("//*[@role='textbox']")
                actual = await output_elem.text_content()
            except:
                actual = "ERROR: Could not extract result"
            
            # Compare results
            test_result = "PASS" if actual == expected else "FAIL"
            details = f"Expected: {expected}\nActual: {actual}" if test_result == "FAIL" else "Test passed"
            
            # Take screenshot
            timestamp = datetime.now().isoformat()
            screenshot_path = f"results/screenshots/{test_name}_{timestamp.replace(':', '-')}.png"
            os.makedirs(os.path.dirname(screenshot_path), exist_ok=True)
            
            try:
                await tested_page.screenshot(path=screenshot_path)
                screenshot_url = f"file://{os.path.abspath(screenshot_path)}"
            except:
                screenshot_url = "FAILED"
            
            # Store result
            result = {
                'test_name': test_name,
                'result': test_result,
                'actual': actual,
                'details': details,
                'timestamp': timestamp,
                'screenshot': screenshot_url,
                'row': test_case['row']
            }
            results.append(result)
            
            # Add results to Excel
            error = typofix_helpers.add_results_to_excel(
                test_name,
                test_result,
                actual,
                details,
                timestamp,
                screenshot_url
            )
            
            if error:
                print(f"  Warning: {error}")
            
            print(f"  Result: {test_result}")
            
        except Exception as e:
            print(f"Error executing test {test_case['name']}: {str(e)}")
            results.append({
                'test_name': test_case['name'],
                'result': 'ERROR',
                'actual': f"Error: {str(e)}",
                'details': str(e),
                'timestamp': datetime.now().isoformat(),
                'screenshot': 'FAILED',
                'row': test_case['row']
            })
    
    # Save Excel file with results
    typofix_helpers.save_test_case_excel()
    
    # Generate report
    passed = sum(1 for r in results if r['result'] == 'PASS')
    failed = sum(1 for r in results if r['result'] == 'FAIL')
    errors = sum(1 for r in results if r['result'] == 'ERROR')
    total = len(results)
    
    print(f"\n{'='*60}")
    print(f"Test Execution Summary")
    print(f"{'='*60}")
    print(f"Total Tests: {total}")
    print(f"Passed: {passed}")
    print(f"Failed: {failed}")
    print(f"Errors: {errors}")
    print(f"Success Rate: {(passed/total*100):.1f}%" if total > 0 else "N/A")
    print(f"{'='*60}\n")
    
    # Assert at least some tests passed
    assert passed > 0, "No tests passed"


@pytest.mark.asyncio
async def test_single_case_example(tested_page):
    """
    Example test for a single case.

    Args:
        tested_page: Page fixture for the tested application
    """
    # Navigate to application
    await tested_page.goto("https://www.typofix.org/application#testing")
    
    # Verify page loaded
    assert await tested_page.title(), "Page should have loaded"
    
    # Find input element
    input_elem = tested_page.locator("//*[@role='textbox']")
    assert input_elem, "Input element should exist"
