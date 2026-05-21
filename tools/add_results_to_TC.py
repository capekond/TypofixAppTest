import shutil
import xml.etree.ElementTree as Et
import argparse
import pandas as pd
from openpyxl import load_workbook
import os
from pathlib import Path

class Common:
    def __init__(self):

        # !!!
        dr =  Path(__file__).parent.parent
        self.RESOURCES_DIR = os.path.join(dr, 'tests',  'resources')

        self.TEST_DATA_DIR = os.path.join(self.RESOURCES_DIR, "test_data")
        self.PCX_DIR = os.path.join(self.TEST_DATA_DIR, "pcx_dir")

        # !!!!
        self.REPORT_DIR = os.path.join(dr, 'results')

        self.DATA_STORE_FILE = os.path.join(self.TEST_DATA_DIR, "DataStore.xlsx")
        self.TEST_CASES_FILE = os.path.join(self.TEST_DATA_DIR, "TestCases.xlsx")
        self.LANGUAGES_FILE = os.path.join(self.TEST_DATA_DIR, 'references', '_list.csv')
        self.REPORT_FILE = os.path.join(self.REPORT_DIR,'output.xml')

        self.HTML_TAGS = ['<br>', '<p>', '<span>']
        self.PATTERN = "_pattern"
        self.CLEAN_CHAR = '_'

    def clean_up_text(self, txt: str) -> str:
        res = ""
        for t in txt:
            res += self.CLEAN_CHAR if t.isspace() or not (t.isalnum()) else t
        return res

class Helpers(Common):
    def __init__(self):
        super().__init__()

    def get_args(self):
        parser = argparse.ArgumentParser()
        parser.add_argument("-p", "--no_pictures", action='store_true', help="Disable screenshot links")
        parser.add_argument("-f", "--update_file", default=self.TEST_CASES_FILE, help=f"Changed Excel file. Implicit value {self.TEST_CASES_FILE}")
        return parser.parse_args()

    @staticmethod
    def get_test_names_from_tc() -> list[str]:
        df = pd.read_excel(p.update_file, sheet_name=0)
        return df['*** Test Cases ***'].tolist()

    def get_results(self, t_names: list[str]) -> list[list[str]]:
        res = []
        to_report = Et.parse(self.REPORT_FILE).getroot()
        for test_name in t_names:
            try:
                rs = []
                xlm_test_name = f"./suite/suite/suite/suite/test[@name='{test_name}']"
                status = to_report.find(xlm_test_name)
                rs.append(status.find('status').attrib['status'])
                if rs[0] == 'FAIL':
                    xlm_error_msg = f"{xlm_test_name}/kw/kw[@name='Element Text Should Be']/msg"
                    msg = to_report.findall(xlm_error_msg)
                    info = msg[3].text.strip().replace('\n', '')
                    rs.append(' '.join(info.split()).replace("'//*[@role=\"textbox\"]'", ''))
                else:
                    rs.append('')
                time_stamp = status.find('status').attrib['start']
                rs.append(time_stamp)
                xml_picture = f"{xlm_test_name}/kw/kw[@name='Capture Element Screenshot']/msg"
                pic = to_report.findall(xml_picture)
                source_pic = self.grep_xml_attribute(pic[0].text,'src', str(m.REPORT_DIR) + os.sep)
                name_pic = self.move_pictures_to_test_data(source_pic, test_name, time_stamp)
                rs.append(name_pic)
                res.append(rs)
            except AttributeError as e:
                print(f"For test '{test_name}' missing data in {self.REPORT_FILE}")
                print(e)
                exit(1)
        return res

    @staticmethod
    def grep_xml_attribute(source:str, att: str, prefix="") -> str:
        att = f'{att}="'
        try:
            source = source[source.index(att) + len(att):]
            source = source[:source.index('"')]
        except ValueError as e:
            print(f"Attribute {att} not found.\n{e}")
        return prefix  + source

    def move_pictures_to_test_data(self,source_file, test_name: str, test_time: str ) -> str:
        if not os.path.exists(self.PCX_DIR):
            os.makedirs(self.PCX_DIR)
        source  = os.path.join(self.REPORT_DIR, source_file )
        target = os.path.join(self.PCX_DIR, test_name + self.clean_up_text(test_time) + ".png" )
        shutil.copyfile(source, target)
        return str(target)

    @staticmethod
    def write_test_names_from_tc(t_results: list[list[str]]) -> None:
        wb = load_workbook(p.update_file)
        sh = wb.worksheets[0]
        for rid, r in enumerate(t_results):
            for cid, c in enumerate(r[:-1]):
                sh.cell(2 + rid, 9 + cid, c)
            link = '=HYPERLINK("{}", "{}")'.format(t_results[rid][3], "LINK")
            sh.cell(2  + rid, 8 + len(t_results),link)
        wb.save(p.update_file)
        wb.close()

if __name__ == "__main__":
    m = Helpers()
    p = m.get_args()
    test_names = m.get_test_names_from_tc()
    print(f"Results added to file {p.update_file} from {m.REPORT_FILE}")
    print(f"\nSelected {len(test_names)} test cases:")
    has_pcx = "NOT " if p.no_pictures else ""
    print(*test_names, sep='\n')
    print(f"\nLink to screenshots is {has_pcx}added.")
    test_results = m.get_results(test_names)
    m.write_test_names_from_tc(test_results)
