import argparse
import os
import re
import sys
from datetime import datetime
from shlex import split

import pandas as pd
from docutils.nodes import target
from openpyxl import load_workbook
from pathlib import Path

from sqlalchemy.sql.functions import concat


class Helpers:
    def __init__(self):
        self.RESOURCES_DIR =  os.path.join(Path(__file__).parent.parent, "tests", "resources", "test_data")
        self.PATTERN = "_pattern"
    def get_args(self):
        parser = argparse.ArgumentParser()
        store = os.path.join(self.RESOURCES_DIR, "DataStore.xlsx")
        tc = os.path.join(self.RESOURCES_DIR, "TestCases.xlsx")
        generic = parser.add_argument_group('Basic arguments')
        generic.add_argument("-n", "--no_question", action='store_true', help="Disable approval question")
        generic.add_argument("-i", "--input", default=store, help=f"Source Excel file. Implicit value {store}")
        generic.add_argument("-o", "--output", default=tc, help=f"Target Excel file. Implicit value {tc}")
        return parser.parse_args()

    def get_sheets(self, source, target):
        wbs = load_workbook(source)
        last = wbs.worksheets[len(wbs.worksheets)-1]
        print(f"Selected source worksheet {last.title}")
        wbt = load_workbook(target)
        first = wbt.copy_worksheet(wbt[self.PATTERN])
        first.title = "tc_A"
        wbt.move_sheet(first, offset=- (len(wbt.worksheets)-1))
        print(f"Selected target worksheet {first.title}")
        return last, first
if __name__ == "__main__":
    m = Helpers()
    p = m.get_args()
    print(f"Content of last list from {p.input} will be formatted and transferred to {p.output} ")
    print("For more info let try --help")
    if p.no_question or input('Do you like to proceed the task? [Y/n]  ') == "Y":
        s, t = m.get_sheets(p.input, p.output)
        rows = [  [{"name" : s.cell(i, 1).value + ". " + s.cell(i, 2).value },
                     {"before": s.cell(i, 3).value},
                     {"after": s.cell(i, 4).value},
                     {"lang": s.cell(i, 5).value},
                     ] for i in range(2, s.max_row + 1)]
        rows_final = []
        for row in rows:
            langs = list(row[3].values())[0]
            for lng in langs.split(","):
                rows_final.append([ (list(row[0].values())[0]).strip() + " [" + lng.strip() + "]",
                                   lng,
                                   list(row[1].values())[0],
                                   list(row[2].values())[0]])
        target_col_index = (1,4,5,6)
        for r_i, row_final in enumerate(rows_final):
            for c_prep, cell_prep in enumerate(row_final):
                cell_final = (str(cell_prep)).strip() if cell_prep else ""
                t.cell(r_i + 2, target_col_index[c_prep], cell_final)
        t.parent.save(p.output)
    else:
        print("Disapproved by user ")

