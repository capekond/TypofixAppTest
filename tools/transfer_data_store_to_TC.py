import argparse
from openpyxl import load_workbook
from pathlib import Path
import os

class Common:
    def __init__(self):

        # !!!
        dr = Path(__file__).parent.parent
        self.RESOURCES_DIR = os.path.join(dr, 'tests', 'resources')

        self.TEST_DATA_DIR = os.path.join(self.RESOURCES_DIR, "test_data")
        self.PCX_DIR = os.path.join(self.TEST_DATA_DIR, "pcx_dir")

        # !!!!
        self.REPORT_DIR = os.path.join(dr, 'results')

        self.DATA_STORE_FILE = os.path.join(self.TEST_DATA_DIR, "DataStore.xlsx")
        self.TEST_CASES_FILE = os.path.join(self.TEST_DATA_DIR, "TestCases.xlsx")
        self.LANGUAGES_FILE = os.path.join(self.TEST_DATA_DIR, 'references', '_list.csv')
        self.REPORT_FILE = os.path.join(self.REPORT_DIR,'output.xml')

        self.HTML_TAGS = ['<br>', '<p>', '<span>']
        self.PATTERN = "_pattern"
        self.CLEAN_CHAR = '_'

    def clean_up_text(self, txt: str) -> str:
        res = ""
        for t in txt:
            res += self.CLEAN_CHAR if t.isspace() or not (t.isalnum()) else t
        return res


class Helpers(Common):
    def __init__(self):
        super().__init__()
    def get_args(self):
        parser = argparse.ArgumentParser()
        parser.add_argument("-n", "--no_question", action='store_true', help="Disable approval question")
        parser.add_argument("-i", "--input", default=self.DATA_STORE_FILE, help=f"Source Excel file. Implicit value {self.DATA_STORE_FILE}")
        parser.add_argument("-o", "--output", default=self.TEST_CASES_FILE, help=f"Target Excel file. Implicit value {self.TEST_CASES_FILE}")
        return parser.parse_args()

    def get_sheets(self, source, target):
        wbs = load_workbook(source)
        last = wbs.worksheets[len(wbs.worksheets)-1]
        print(f"Selected source worksheet {last.title}")
        wbt = load_workbook(target)
        first = wbt.copy_worksheet(wbt[self.PATTERN])
        first.title = "tc_A"
        wbt.move_sheet(first, offset=- (len(wbt.worksheets)-1))
        print(f"Selected target worksheet {first.title}")
        return last, first

if __name__ == "__main__":
    m = Helpers()
    p = m.get_args()
    print(f"Content of last list from {p.input} will be formatted and transferred to {p.output} ")
    print("For more info let try --help")
    if p.no_question or input('Do you like to proceed the task? [Y/n]  ') == "Y":
        s, t = m.get_sheets(p.input, p.output)
        rows = [  [{"name" : s.cell(i, 1).value + ". " + s.cell(i, 2).value },
                     {"before": s.cell(i, 3).value},
                     {"after": s.cell(i, 4).value},
                     {"lang": s.cell(i, 5).value},
                     ] for i in range(2, s.max_row + 1)]
        rows_final = []
        for row in rows:
            langs = list(row[3].values())[0]
            for lng in langs.split(","):
                t_name = m.clean_up_text((list(row[0].values())[0]).strip() + lng.strip())
                rows_final.append([ t_name,
                                    lng,
                                    list(row[1].values())[0],
                                    list(row[2].values())[0]])
        target_col_index = (1,4,5,6)
        for r_i, row_final in enumerate(rows_final):
            for c_prep, cell_prep in enumerate(row_final):
                cell_final = (str(cell_prep)).strip() if cell_prep else ""
                t.cell(r_i + 2, target_col_index[c_prep], cell_final)
        t.parent.save(p.output)
    else:
        print("Disapproved by user ")